#!/usr/bin/python3

from functools import reduce
import re
import sys

import click
import psycopg2
import psycopg2.extras
from openpyxl import load_workbook

GET_ERA_BY_YEAR_QUERY = '''
with tt_era as (
  select e.era,
         e.year + es.weight as approximate_year
    from tb_era e
    join tb_era_specifier es
      on e.era_specifier = es.era_specifier
)
   select distinct on ( s )
          s,
          e.era
     from generate_series(1970, 2029) s
left join tt_era e
       on e.approximate_year = s
       or abs(s - e.approximate_year) <= 1
 order by s,
          e.approximate_year desc
'''

GET_ERA_BY_READABLE_VALUE_QUERY = '''
select e.era,
       lower(es.label || ' ' || e.year || 's') as human_readable
  from tb_era e
  join tb_era_specifier es
    on e.era_specifier = es.era_specifier
'''

INSERT_AESTHETIC_QUERY = '''
insert into tb_aesthetic (
    name,
    url_slug,
    start_era,
    end_era,
    description,
    media_source_url,
    creator,
    modifier
) values (
    %(name)s,
    %(url_slug)s,
    %(start_year)s,
    %(end_year)s,
    %(description)s,
    %(media_source_url)s,
    0,
    0
) returning aesthetic
'''

INSERT_WEBSITE_QUERY = '''
insert into tb_aesthetic_website as aw (
    aesthetic,
    url,
    website_type,
    creator,
    modifier
)
   select %(aesthetic)s,
          %(url)s,
          wt.website_type,
          0,
          0
     from tb_website_type wt
    where regexp_replace(wt.label, '[^a-zA-Z0-9]', '', 'g') ilike %(website_type_label)s || '%%'
       on conflict (aesthetic, url)
       do update
      set url = EXCLUDED.url
returning aesthetic_website
'''

INSERT_AESTHETIC_RELATIONSHIP_QUERY = '''
with tt_aesthetic_relationship as (
    select from_a.aesthetic as from_aesthetic,
           to_a.aesthetic   as to_aesthetic,
           %(description)s  as description
      from tb_aesthetic from_a,
           tb_aesthetic to_a
     where from_a.name = %(from_aesthetic_name)s
       and to_a.name = %(to_aesthetic_name)s
     union all
    select to_a.aesthetic          as from_aesthetic,
           from_a.aesthetic        as to_aesthetic,
           %(reverse_description)s as description
      from tb_aesthetic from_a,
           tb_aesthetic to_a
     where from_a.name = %(from_aesthetic_name)s
       and to_a.name = %(to_aesthetic_name)s
)
insert into tb_aesthetic_relationship as ar (
    from_aesthetic,
    to_aesthetic,
    description,
    creator,
    modifier
)
   select ttar.from_aesthetic,
          ttar.to_aesthetic,
          ttar.description,
          0,
          0
     from tt_aesthetic_relationship ttar
       on conflict (from_aesthetic, to_aesthetic) do update
      set description = excluded.description
returning aesthetic_relationship
'''

INSERT_MEDIA_CREATOR_QUERY = '''
insert into tb_media_creator (
    name
) values (
    %(name)s
)
returning media_creator
'''

INSERT_MEDIA_QUERY = '''
insert into tb_aesthetic_media (
    aesthetic,
    url,
    preview_image_url,
    label,
    description,
    media_creator,
    year,
    creator,
    modifier
)
   select a.aesthetic,
          %(url)s,
          %(preview_image_url)s,
          %(label)s,
          %(description)s,
          %(media_creator)s,
          %(year)s,
          0,
          0
     from tb_aesthetic a
    where a.name = %(aesthetic_name)s
returning aesthetic_media
'''

aesthetic_names = {}
db_handle = None
has_error = False
use_default_values = False
readable_value_era_table = {}
year_era_table = {}


def build_year_era_table():
    global db_handle

    year_era_rows = query(GET_ERA_BY_YEAR_QUERY)

    def reducer(table, row):
        table[str(row['s'])] = row['era']
        return table

    return reduce(reducer, year_era_rows, {})


def build_readable_value_era_table():
    global db_handle

    readable_value_era_rows = query(GET_ERA_BY_READABLE_VALUE_QUERY)

    def reducer(table, row):
        table[row['human_readable']] = row['era']
        return table

    return reduce(reducer, readable_value_era_rows, {})


def query(query, **kwargs):
    global db_handle

    rval = []

    with db_handle.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cursor:
        cursor.execute(query, kwargs)
        rval = list(map(dict, cursor.fetchall()))

    return rval


def comma_split(value):
    return list(filter(lambda v: v != '', map(lambda v: v.strip(), value.split(',')))) if value else []


def parse_header_row(worksheet):
    return {
        col.value: col.col_idx - 1
        for col in worksheet[1]
    }


def parse_aesthetic_row(cells, headers):
    global readable_value_era_table, year_era_table
    rval = {}

    for key in ('name', 'description'):
        raw_value = cells[headers[key]].value
        rval[key] = str(raw_value).strip() if raw_value else None

    for key in ('start_year', 'end_year'):
        raw_value = cells[headers[key]].value
        era = None

        if raw_value:
            value_key = str(raw_value).strip().lower()
            era = year_era_table.get(value_key, readable_value_era_table.get(value_key))

            if value_key != 'present' and not era:
                print(
                    f'WARNING: Could not parse value "{raw_value}". (Row: {cells[0].row})')

        rval[key] = era

    rval['websites'] = []

    websites_arena = comma_split(cells[headers['website_arena']].value)
    media_source_url = None

    if websites_arena:
        arena_slug = websites_arena[0].strip().split(
            '/').pop()

        media_source_url = 'https://api.are.na/v2/channels/' + arena_slug

        for website_arena in websites_arena:
            rval['websites'].append({
                'url': website_arena,
                'website_type_label': 'arena'
            })

    rval['media_source_url'] = media_source_url

    for key in ('website_facebook', 'website_twitter', 'website_tumblr'):
        websites = comma_split(cells[headers[key]].value)

        for website in websites:
            if website:
                website_type = key.split('_', 1)

                rval['websites'].append({
                    'url': website,
                    'website_type_label': website_type[1]
                })

    rval['url_slug'] = re.sub(r'\s+', '-', re.sub(r'[^a-zA-Z0-9-\s]', '', rval['name'],
                                                  re.IGNORECASE)).lower()

    return rval


def parse_timeline_row(cells, headers):
    rval = {}

    rval['aesthetic_name'] = cells[headers['aesthetic']].value
    rval['media_creator_name'] = cells[headers['creator']].value

    for key in ('url', 'preview_image_url', 'label', 'description'):
        raw_value = cells[headers[key]].value
        rval[key] = str(raw_value).strip() if raw_value else None

    year = cells[headers['year']].value
    rval['year'] = int(year) if year else None

    return rval


def parse_similarity_row(cells, headers):
    rval = {}

    raw_from_aesthetic = cells[headers['from_aesthetic']].value
    raw_to_aesthetic = cells[headers['to_aesthetic']].value

    rval['from_aesthetic_name'] = str(raw_from_aesthetic).strip(
    ) if raw_from_aesthetic else None

    rval['to_aesthetic_name'] = str(raw_to_aesthetic).strip(
    ) if raw_to_aesthetic else None

    for key in ('description', 'reverse_description'):
        raw_value = cells[headers[key]].value
        rval[key] = str(raw_value).strip() if raw_value else None

    return rval


def process_aesthetics_sheet(worksheet):
    global aesthetic_names, db_handle, has_error, use_default_values

    headers = parse_header_row(worksheet)
    url_slugs = {}

    for cells in worksheet.iter_rows(min_row=2):
        parsed_row = parse_aesthetic_row(cells, headers)
        skip_row = False

        if not parsed_row['name']:
            print(
                f'ERROR: "name" is required. (Row: {cells[0].row})')

            skip_row = True

        if not parsed_row['description']:
            if use_default_values:
                parsed_row['description'] = 'This aesthetic is still being researched. Please check back later!'
            else:
                print(
                    f'ERROR: "description" is required. (Row: {cells[0].row})')
                skip_row = True

        name = parsed_row['name']

        if aesthetic_names.get(name):
            print(
                f'ERROR: Aesthetic name "{name}" is already in use. (Row: {cells[0].row})')

            skip_row = True
        else:
            aesthetic_names[name] = True

        url_slug = parsed_row['url_slug']

        if url_slugs.get(url_slug):
            print(
                f'ERROR: URL slug "{url_slug}" is already in use. (Row: {cells[0].row})')

            skip_row = True
        else:
            url_slugs[url_slug] = True

        if skip_row:
            has_error = True
            continue

        aesthetic_row = {
            'name': name,
            'url_slug': url_slug,
            'start_year': parsed_row['start_year'],
            'end_year': parsed_row['end_year'],
            'description': f'<p>{parsed_row["description"]}</p>',
            'media_source_url': parsed_row['media_source_url'],
        }

        pk_aesthetic = query(
            INSERT_AESTHETIC_QUERY, **aesthetic_row)[0]['aesthetic']

        for website in parsed_row['websites']:
            query(INSERT_WEBSITE_QUERY, **
                  {'aesthetic': pk_aesthetic, **website})


def process_timeline_sheet(worksheet):
    global aesthetic_names, db_handle

    if worksheet.min_row < 2:
        return

    headers = parse_header_row(worksheet)
    media_creators = {}

    for cells in worksheet.iter_rows(min_row=2):
        parsed_row = parse_timeline_row(cells, headers)
        skip_row = False

        for required_column in ('aesthetic_name', 'url', 'preview_image_url', 'label', 'description', 'year'):
            if not parsed_row[required_column]:
                print(
                    f'ERROR: "{required_column}" is required. (Row: {cells[0].row})')
                skip_row = True

        year = parsed_row['year']

        try:
            int(year)
        except ValueError:
            print(f'ERROR: "year" must be numeric. (Row: {cells[0].row})')
            skip_row = True

        aesthetic_name = parsed_row['aesthetic_name']

        if not aesthetic_names.get(aesthetic_name):
            print(
                f'ERROR: Aesthetic "{aesthetic_name}" does not exist. Please check the aesthetics worksheet and try again. (Row: {cells[0].row})')

            skip_row = True

        if skip_row:
            has_error = True
            continue

        creator_name = parsed_row.get('media_creator_name')

        if creator_name and not media_creators.get(creator_name):
            media_creators[creator_name] = query(
                INSERT_MEDIA_CREATOR_QUERY, name=creator_name)[0]['media_creator']

        media_row = {
            'aesthetic_name': aesthetic_name,
            'url': parsed_row['url'],
            'preview_image_url': parsed_row['preview_image_url'],
            'label': parsed_row['label'],
            'description': parsed_row['description'],
            'media_creator': media_creators[creator_name] if creator_name else None,
            'year': parsed_row['year'],
        }

        query(INSERT_MEDIA_QUERY, **media_row)


def process_similarity_sheet(worksheet):
    global aesthetic_names, db_handle

    headers = parse_header_row(worksheet)
    aesthetics = {}
    aesthetic_relationships = {}

    for cells in worksheet.iter_rows(min_row=2):
        parsed_row = parse_similarity_row(cells, headers)
        skip_row = False

        for required_column in ('from_aesthetic_name', 'to_aesthetic_name'):
            if not parsed_row[required_column]:
                print(
                    f'ERROR: "{required_column}" is required. (Row: {cells[0].row})')
                skip_row = True

        from_aesthetic = parsed_row['from_aesthetic_name']
        to_aesthetic = parsed_row['to_aesthetic_name']

        if not aesthetic_names.get(from_aesthetic):
            print(
                f'ERROR: Aesthetic "{from_aesthetic}" does not exist. Please check the aesthetics worksheet and try again. (Row: {cells[0].row})')

            skip_row = True

        if not aesthetic_names.get(to_aesthetic):
            print(
                f'ERROR: Aesthetic "{to_aesthetic}" does not exist. Please check the aesthetics worksheet and try again. (Row: {cells[0].row})')

            skip_row = True

        if aesthetic_relationships.get(from_aesthetic, {}).get(to_aesthetic) or aesthetic_relationships.get(to_aesthetic, {}).get(from_aesthetic):
            print(
                f'WARNING: Relationship between "{from_aesthetic}" to "{to_aesthetic}" already defined. (Row: {cells[0].row})')

        if not aesthetic_relationships.get(from_aesthetic, {}):
            aesthetic_relationships[from_aesthetic] = {}

        if not aesthetic_relationships.get(to_aesthetic, {}):
            aesthetic_relationships[to_aesthetic] = {}

        aesthetic_relationships[from_aesthetic][to_aesthetic] = True
        aesthetic_relationships[to_aesthetic][from_aesthetic] = True

        if skip_row:
            has_error = True
            continue

        query(INSERT_AESTHETIC_RELATIONSHIP_QUERY, **parsed_row)


@click.command()
@click.option('-h', '--host', default='localhost', help='database server host or socket directory')
@click.option('-U', '--username', default='cari', help='database user name')
@click.option('-d', '--dbname', default='cari', help='database name to connect to')
@click.option('-p', '--port', default='5432', help='database server port')
@click.option('-W', '--password', help='force password prompt')
@click.option('--use-defaults', is_flag=True, help='use sensible default values for empty cells where appropriate')
@click.argument('datafile', type=click.Path(exists=True))
def main(host, username, dbname, port, password, use_defaults, datafile):
    global db_handle, has_error, readable_value_era_table, use_default_values, year_era_table

    use_default_values = use_defaults
    workbook = load_workbook(filename=datafile, data_only=True)

    psql_connection_args = {
        'dbname': dbname,
        'user': username,
        'host': host,
        'port': port,
    }

    if(password):
        psql_connection_args['password'] = password

    db_handle = psycopg2.connect(**psql_connection_args)

    worksheets = {
        workbook.sheetnames[i]: i
        for i in range(len(workbook.sheetnames))
    }

    year_era_table = build_year_era_table()
    readable_value_era_table = build_readable_value_era_table()

    print('Processing aesthetics worksheet...')
    workbook.active = worksheets['aesthetics']
    process_aesthetics_sheet(workbook.active)

    print('Processing timeline worksheet...')
    workbook.active = worksheets['timeline']
    process_timeline_sheet(workbook.active)

    print('Processing similarity worksheet...')
    workbook.active = worksheets['similarity']
    process_similarity_sheet(workbook.active)

    if has_error:
        print('Workbook has errors. Please correct and try again.')
        db_handle.rollback()
        sys.exit(-1)

    db_handle.commit()
    print('Done')
    sys.exit(0)


if __name__ == '__main__':
    main()
